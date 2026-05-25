"""openpyxl tracker — Pipeline, Applied, and Log sheets."""
from __future__ import annotations

import datetime
from datetime import timezone
from pathlib import Path
from typing import Optional

from loguru import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.views import SheetView, Pane

from agent.config import Settings
from agent.search.base import JobListing

# Column definitions for Pipeline and Applied sheets
PIPELINE_HEADERS = [
    "Rank", "Company", "Role Title", "Location", "Source",
    "Score", "Tier", "Role Family", "Fit Summary",
    "Apply Link", "CV Ready", "Cover Letter Ready",
    "Status", "Applied Date", "Notes", "First Seen", "Last Updated",
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="EAF0FB")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")

COL_WIDTHS = {
    "A": 5,  "B": 22, "C": 30, "D": 16, "E": 9,
    "F": 7,  "G": 9,  "H": 14, "I": 50, "J": 35,
    "K": 10, "L": 14, "M": 14, "N": 13, "O": 40,
    "P": 20, "Q": 20,
}


def _now_iso() -> str:
    return datetime.datetime.now(timezone.utc).isoformat(timespec="seconds")


class TrackerWorkbook:
    def __init__(self, settings: Settings) -> None:
        self.path: Path = settings.tracker_path / "george_emil_job_tracker.xlsx"
        self.wb: Workbook | None = None
        self._pipeline = None
        self._applied  = None
        self._log      = None

    def load_or_create(self) -> None:
        if self.path.exists():
            self.wb = load_workbook(self.path)
            logger.info(f"Tracker loaded: {self.path}")
        else:
            self.wb = Workbook()
            self._bootstrap()
            logger.info(f"Tracker created: {self.path}")
        self._pipeline = self.wb["Pipeline"]
        self._applied  = self.wb["Applied"]
        self._log      = self.wb["Log"]

    def _bootstrap(self) -> None:
        """Create the three sheets with headers and formatting."""
        ws_pipe = self.wb.active
        ws_pipe.title = "Pipeline"
        self._write_headers(ws_pipe, PIPELINE_HEADERS)

        ws_applied = self.wb.create_sheet("Applied")
        self._write_headers(ws_applied, PIPELINE_HEADERS)

        ws_log = self.wb.create_sheet("Log")
        self._write_headers(ws_log, ["Timestamp", "Event", "Detail"])
        ws_log.column_dimensions["A"].width = 22
        ws_log.column_dimensions["B"].width = 30
        ws_log.column_dimensions["C"].width = 80

    def _write_headers(self, ws, headers: list[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            letter = get_column_letter(col_idx)
            if letter in COL_WIDTHS:
                ws.column_dimensions[letter].width = COL_WIDTHS[letter]
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "B2"
        if len(headers) > 1:
            ws.auto_filter.ref = ws.dimensions

    def _row_index_for_slug(self, slug: str) -> Optional[int]:
        """Return 1-based row index in Pipeline sheet, or None."""
        for row in self._pipeline.iter_rows(min_row=2, values_only=False):
            notes_cell = row[14]   # col O = Notes (0-indexed 14)
            slug_cell  = row[14]   # we store slug in Notes col for lookup
            # Slug is stored as a hidden marker in column O or we match via company+title
            # We use a hidden slug marker format: "slug:<value>" appended to Notes
            if notes_cell.value and f"slug:{slug}" in str(notes_cell.value):
                return notes_cell.row
        return None

    def upsert_role(self, listing: JobListing, score_result: dict) -> None:
        """Insert new role or update existing one in Pipeline sheet."""
        existing_row = self._row_index_for_slug(listing.slug)
        now = _now_iso()

        if existing_row:
            ws = self._pipeline
            ws.cell(row=existing_row, column=7).value  = score_result.get("tier", "")
            ws.cell(row=existing_row, column=6).value  = score_result.get("score", 0)
            ws.cell(row=existing_row, column=9).value  = score_result.get("fit_summary", "")
            ws.cell(row=existing_row, column=17).value = now
            logger.debug(f"Tracker: updated '{listing.title}' @ {listing.company}")
        else:
            ws = self._pipeline
            next_row = ws.max_row + 1
            rank = next_row - 1

            row_data = [
                rank,
                listing.company,
                listing.title,
                listing.location,
                listing.source,
                score_result.get("score", 0),
                score_result.get("tier", ""),
                score_result.get("role_family", ""),
                score_result.get("fit_summary", ""),
                listing.apply_url,
                "No",   # CV Ready
                "No",   # Cover Letter Ready
                "Not Applied",
                "",     # Applied Date
                f"slug:{listing.slug}",  # Notes (slug marker)
                listing.fetched_at,
                now,
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                # Alternating row fill
                if next_row % 2 == 0:
                    cell.fill = ALT_FILL
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 9))

            # Highlight Apply Link yellow if Not Applied
            j_cell = ws.cell(row=next_row, column=10)
            j_cell.fill = YELLOW_FILL
            if listing.apply_url:
                j_cell.hyperlink = listing.apply_url
                j_cell.font = Font(color="0000FF", underline="single")

            self.append_log("new_role", f"{listing.company} — {listing.title} [{listing.source}] score={score_result.get('score')}")
            logger.debug(f"Tracker: inserted '{listing.title}' @ {listing.company}")

    def set_status(self, slug: str, status: str) -> None:
        row = self._row_index_for_slug(slug)
        if row:
            self._pipeline.cell(row=row, column=13).value = status
            self._pipeline.cell(row=row, column=17).value = _now_iso()

    def mark_draft(self, slug: str) -> None:
        """Role has generated artifacts awaiting human review."""
        self.set_status(slug, "Draft")

    def mark_ready_for_apply(self, slug: str) -> None:
        """Human approved artifacts — ready to apply."""
        self.set_status(slug, "Ready")

    def get_row_by_slug(self, slug: str) -> list | None:
        row_idx = self._row_index_for_slug(slug)
        if not row_idx:
            return None
        return list(
            self._pipeline.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
        )[0]

    def mark_applied(self, slug: str, applied_date: str) -> None:
        row = self._row_index_for_slug(slug)
        if not row:
            logger.warning(f"Tracker.mark_applied: slug not found: {slug}")
            return
        ws = self._pipeline
        status_cell = ws.cell(row=row, column=13)
        if status_cell.value == "Applied":
            logger.warning(f"Tracker: already applied: {slug}")
            return
        status_cell.value = "Applied"
        ws.cell(row=row, column=14).value = applied_date
        ws.cell(row=row, column=17).value = _now_iso()
        # Remove yellow highlight
        ws.cell(row=row, column=10).fill = PatternFill()
        self.append_log("applied", f"slug={slug} date={applied_date}")

    def mark_cv_ready(self, slug: str) -> None:
        row = self._row_index_for_slug(slug)
        if row:
            self._pipeline.cell(row=row, column=11).value = "Yes"
            self._pipeline.cell(row=row, column=17).value = _now_iso()
            self.append_log("cv_ready", f"slug={slug}")

    def mark_letter_ready(self, slug: str) -> None:
        row = self._row_index_for_slug(slug)
        if row:
            self._pipeline.cell(row=row, column=12).value = "Yes"
            self._pipeline.cell(row=row, column=17).value = _now_iso()
            self.append_log("letter_ready", f"slug={slug}")

    def get_all_slugs(self) -> set[str]:
        slugs: set[str] = set()
        for ws in [self._pipeline, self._applied]:
            if ws is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                notes = str(row[14]) if len(row) > 14 and row[14] else ""
                if "slug:" in notes:
                    slug = notes.split("slug:")[1].split()[0]
                    slugs.add(slug)
        return slugs

    def rerank(self) -> None:
        """Recalculate Rank column by score descending."""
        ws = self._pipeline
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        rows.sort(key=lambda r: (r[5] or 0), reverse=True)
        for i, row_data in enumerate(rows, start=1):
            row_data[0] = i
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=i + 1, column=col_idx).value = value

    def append_log(self, event: str, detail: str) -> None:
        ws = self._log
        if ws is None:
            return
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = _now_iso()
        ws.cell(row=next_row, column=2).value = event
        ws.cell(row=next_row, column=3).value = detail

    def save(self) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(self.path)
        logger.info(f"Tracker saved: {self.path}")
