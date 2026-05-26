"""Import/export helpers between Excel and the SQL tracker."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from agent.config import Settings
from agent.search.base import JobListing, make_slug
from agent.tracker import get_tracker
from agent.tracker.models import RoleRecord
from agent.tracker.workbook import PIPELINE_HEADERS


def _yes(value) -> bool:
    return str(value or "").strip().lower() in {"yes", "true", "1"}


def _slug_from_notes(notes: str, record: RoleRecord) -> str:
    if "slug:" in notes:
        return notes.split("slug:", 1)[1].split()[0]
    listing = JobListing(
        title=record.title,
        company=record.company,
        location=record.location,
        source=record.source or "manual",
        apply_url=record.apply_url,
    )
    return make_slug(listing)


def _record_from_row(row: tuple) -> RoleRecord:
    record = RoleRecord(
        slug="",
        rank=int(row[0] or 0),
        company=str(row[1] or ""),
        title=str(row[2] or ""),
        location=str(row[3] or ""),
        source=str(row[4] or ""),
        score=int(row[5] or 0),
        tier=str(row[6] or ""),
        role_family=str(row[7] or ""),
        fit_summary=str(row[8] or ""),
        apply_url=str(row[9] or ""),
        cv_ready=_yes(row[10]),
        letter_ready=_yes(row[11]),
        status=str(row[12] or "Not Applied"),
        applied_date=str(row[13]) if row[13] else None,
        first_seen=str(row[15]) if len(row) > 15 and row[15] else None,
        last_updated=str(row[16]) if len(row) > 16 and row[16] else None,
        scoring_status=str(row[17]) if len(row) > 17 and row[17] else "",
        artifact_status=str(row[18]) if len(row) > 18 and row[18] else "none",
        ats_keywords=str(row[19]) if len(row) > 19 and row[19] else "",
    )
    record.slug = _slug_from_notes(str(row[14] or ""), record)
    return record


def import_tracker(settings: Settings, source: Path | None = None) -> int:
    source = source or (settings.tracker_path / "george_emil_job_tracker.xlsx")
    tracker = get_tracker(settings)
    tracker.load_or_create()

    wb = load_workbook(source, data_only=True)
    ws = wb["Pipeline"]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or not row[2]:
            continue
        tracker.upsert_record(_record_from_row(row))
        count += 1

    if "Log" in wb.sheetnames:
        existing = {(e.event, e.detail) for e in tracker.list_events()}
        for row in wb["Log"].iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                key = (str(row[1] or ""), str(row[2] or ""))
                if key in existing:
                    continue
                tracker.append_log(key[0], key[1])
                existing.add(key)

    tracker.rerank()
    return count


def export_tracker(settings: Settings, output: Path | None = None) -> Path:
    output = output or (settings.tracker_path / "george_emil_job_tracker.xlsx")
    tracker = get_tracker(settings)
    tracker.load_or_create()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pipeline"
    ws.append(PIPELINE_HEADERS)
    applied = wb.create_sheet("Applied")
    applied.append(PIPELINE_HEADERS)
    log = wb.create_sheet("Log")
    log.append(["Timestamp", "Event", "Detail"])

    for record in tracker.list_pipeline_rows(include_applied=True):
        target = applied if record.status == "Applied" else ws
        target.append(
            [
                record.rank,
                record.company,
                record.title,
                record.location,
                record.source,
                record.score,
                record.tier,
                record.role_family,
                record.fit_summary,
                record.apply_url,
                "Yes" if record.cv_ready else "No",
                "Yes" if record.letter_ready else "No",
                record.status,
                record.applied_date or "",
                f"slug:{record.slug}",
                record.first_seen or "",
                record.last_updated or "",
                record.scoring_status or "",
                record.artifact_status or "none",
                record.ats_keywords or "",
            ]
        )

    for event in tracker.list_events():
        log.append([event.timestamp or "", event.event, event.detail])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output
